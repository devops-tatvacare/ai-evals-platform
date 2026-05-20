"""XLSX → ImportedDataset parser. Shape validation lives in dataset_validator."""
from __future__ import annotations

import io
import zipfile
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.services.orchestration.datasets.dataset_validator import (
    DatasetImportError,
    ImportedDataset,
    MAX_ROWS,
    assemble,
    validate_headers,
    validate_id_strategy,
)


def parse_xlsx(
    raw: bytes,
    *,
    id_strategy: str,
    id_column: Optional[str],
) -> ImportedDataset:
    validate_id_strategy(id_strategy, id_column)

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except (
        InvalidFileException,
        zipfile.BadZipFile,
        KeyError,
        OSError,
        ValueError,
        TypeError,
    ) as exc:
        raise DatasetImportError(
            "file is not a valid .xlsx workbook"
        ) from exc

    worksheet = workbook.active
    if worksheet is None:
        raise DatasetImportError("workbook has no sheets")

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise DatasetImportError("file has no header row")

    columns = [_stringify(c).strip() for c in header_row]
    if any(c == "" for c in columns):
        raise DatasetImportError("file header has a blank column name")
    validate_headers(columns, id_strategy=id_strategy, id_column=id_column)

    rows: list[dict] = []
    for raw_row in rows_iter:
        if all(cell is None or _stringify(cell).strip() == "" for cell in raw_row):
            continue
        if len(rows) >= MAX_ROWS:
            raise DatasetImportError(
                f"row cap exceeded: dataset versions are capped at {MAX_ROWS} rows"
            )
        rows.append({
            columns[i]: _stringify(raw_row[i] if i < len(raw_row) else None).strip()
            for i in range(len(columns))
        })

    return assemble(columns, rows, id_strategy=id_strategy, id_column=id_column)


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)
